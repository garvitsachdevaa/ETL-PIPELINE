import logging
import io
import uuid
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from ingestion.schemas import DocumentObject
from handlers.binary_schema import BinaryDocument, Page, Block, Region

logger = logging.getLogger(__name__)

def handle_xlsx(doc: DocumentObject) -> BinaryDocument:
    """Handle XLSX files with full spreadsheet structure preservation"""
    try:
        # Extract comprehensive spreadsheet content
        content = extract_xlsx_content(doc.raw_bytes)
        
        # Create pages for each worksheet
        pages = []
        for sheet_idx, sheet_data in enumerate(content['worksheets']):
            regions = _create_sheet_regions(sheet_data, sheet_idx)
            
            # Wrap regions in a Block for this worksheet
            blocks = []
            if regions:
                blocks.append(Block(
                    block_id=str(uuid.uuid4()),
                    title=sheet_data['name'],
                    label="worksheet",
                    bbox=[0, 0, 100, 100],
                    regions=regions,
                    raw_text="",  # Could aggregate if needed
                    confidence=0.9,
                    metadata={"engine": "openpyxl", "sheet": sheet_data['name']}
                ))
            
            page = Page(
                page_id=str(uuid.uuid4()),
                page_number=sheet_idx + 1,
                blocks=blocks,
                metadata={
                    "xlsx": True,
                    "sheet_name": sheet_data['name'],
                    "sheet_index": sheet_idx,
                    "dimensions": sheet_data['dimensions'],
                    "has_data": sheet_data['has_data'],
                    "cell_count": sheet_data['cell_count'],
                    "has_formulas": sheet_data['has_formulas'],
                    "has_charts": sheet_data['has_charts']
                }
            )
            pages.append(page)
        
        return BinaryDocument(
            document_id=doc.document_id,
            pages=pages,
            metadata={
                "source_format": doc.detected_format,
                "mime_type": doc.mime_type,
                "extraction_method": "openpyxl",
                "file_size": len(doc.raw_bytes),
                "workbook_stats": content['stats'],
                "total_sheets": len(content['worksheets']),
                "sheet_names": [sheet['name'] for sheet in content['worksheets']]
            }
        )
        
    except Exception as e:
        logger.error(f"Failed to handle XLSX document {doc.document_id}: {e}")
        # Return empty document with error info
        return BinaryDocument(
            document_id=doc.document_id,
            pages=[],
            metadata={
                "source_format": doc.detected_format,
                "mime_type": doc.mime_type,
                "error": str(e),
                "file_size": len(doc.raw_bytes)
            }
        )

def extract_xlsx_content(raw_bytes: bytes) -> Dict[str, Any]:
    """Extract comprehensive content from XLSX file bytes"""
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=False)
        
        worksheets = []
        total_cells = 0
        total_formulas = 0
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            sheet_data = _extract_worksheet_data(worksheet)
            worksheets.append(sheet_data)
            total_cells += sheet_data['cell_count']
            total_formulas += sheet_data['formula_count']
        
        return {
            'worksheets': worksheets,
            'stats': {
                'total_sheets': len(worksheets),
                'total_cells': total_cells,
                'total_formulas': total_formulas,
                'has_named_ranges': len(workbook.defined_names) > 0,
                'sheet_names': workbook.sheetnames
            }
        }
    except Exception as e:
        logger.error(f"Failed to extract XLSX content: {e}")
        raise RuntimeError(f"Failed to extract XLSX content: {str(e)}")

def _extract_worksheet_data(worksheet: Worksheet) -> Dict[str, Any]:
    """Extract structured data from a worksheet"""
    try:
        # Get worksheet dimensions
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            # Empty sheet
            dimensions = "A1:A1"
            has_data = False
        else:
            dimensions = f"{worksheet.min_column}{worksheet.min_row}:{worksheet.max_column}{worksheet.max_row}"
            has_data = True
        
        # Extract cell data with structure preservation
        rows = []
        formula_count = 0
        cell_count = 0
        
        if has_data:
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
                row_data = []
                for col_idx, cell in enumerate(row, 1):
                    cell_info = _extract_cell_info(cell, row_idx, col_idx)
                    row_data.append(cell_info)
                    if cell_info['value'] is not None:
                        cell_count += 1
                    if cell_info['has_formula']:
                        formula_count += 1
                
                # Only include rows with at least one non-empty cell
                if any(cell['value'] is not None for cell in row_data):
                    rows.append({
                        'row_number': row_idx,
                        'cells': row_data
                    })
        
        # Check for charts (basic detection)
        has_charts = len(worksheet._charts) > 0 if hasattr(worksheet, '_charts') else False
        
        # Extract table structures if any
        tables = _extract_table_structures(rows) if rows else []
        
        return {
            'name': worksheet.title,
            'dimensions': dimensions,
            'has_data': has_data,
            'cell_count': cell_count,
            'formula_count': formula_count,
            'has_formulas': formula_count > 0,
            'has_charts': has_charts,
            'rows': rows,
            'tables': tables,
            'row_count': len(rows),
            'column_count': worksheet.max_column if has_data else 0
        }
    except Exception as e:
        logger.warning(f"Failed to extract worksheet data: {e}")
        return {
            'name': worksheet.title,
            'dimensions': "A1:A1",
            'has_data': False,
            'cell_count': 0,
            'formula_count': 0,
            'has_formulas': False,
            'has_charts': False,
            'rows': [],
            'tables': [],
            'row_count': 0,
            'column_count': 0
        }

def _extract_cell_info(cell: Optional[Cell], row_idx: int, col_idx: int) -> Dict[str, Any]:
    """Extract detailed information from a cell"""
    if cell is None or cell.value is None:
        return {
            'coordinate': f"{chr(64 + col_idx)}{row_idx}",
            'value': None,
            'display_value': '',
            'data_type': 'empty',
            'has_formula': False,
            'formula': None,
            'number_format': None
        }
    
    try:
        return {
            'coordinate': cell.coordinate,
            'value': cell.value,
            'display_value': str(cell.value) if cell.value is not None else '',
            'data_type': str(cell.data_type),
            'has_formula': cell.data_type == 'f',
            'formula': getattr(cell, 'formula', None) if cell.data_type == 'f' else None,
            'number_format': cell.number_format
        }
    except Exception:
        return {
            'coordinate': f"{chr(64 + col_idx)}{row_idx}",
            'value': cell.value,
            'display_value': str(cell.value) if cell.value is not None else '',
            'data_type': 'unknown',
            'has_formula': False,
            'formula': None,
            'number_format': None
        }

def _extract_table_structures(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Detect and extract table-like structures from the data"""
    if not rows or len(rows) < 2:
        return []
    
    tables = []
    
    # Simple heuristic: if first row looks like headers and subsequent rows have data
    first_row = rows[0]
    if len(rows) > 1:
        # Check if first row could be headers (non-numeric text values)
        potential_headers = []
        for cell in first_row['cells']:
            if cell['value'] is not None:
                potential_headers.append(cell['display_value'])
            else:
                potential_headers.append('')
        
        if len([h for h in potential_headers if h.strip()]) >= 2:  # At least 2 non-empty headers
            # Extract data rows
            data_rows = []
            for row in rows[1:]:
                data_row = []
                for cell in row['cells']:
                    data_row.append(cell['display_value'])
                data_rows.append(data_row)
            
            tables.append({
                'start_row': first_row['row_number'],
                'end_row': rows[-1]['row_number'],
                'headers': potential_headers,
                'data_rows': data_rows,
                'total_rows': len(rows),
                'total_cols': len(potential_headers)
            })
    
    return tables

def _create_sheet_regions(sheet_data: Dict[str, Any], sheet_idx: int) -> List[Region]:
    """Create regions for different parts of a worksheet"""
    regions = []
    
    if not sheet_data['has_data']:
        # Empty sheet
        regions.append(Region(
            region_id=str(uuid.uuid4()),
            text=f"[Empty worksheet: {sheet_data['name']}]",
            bbox=[0, 0, 100, 100],
            confidence=1.0,
            metadata={
                "engine": "openpyxl",
                "content_type": "empty_sheet",
                "sheet_name": sheet_data['name']
            }
        ))
        return regions
    
    # Create table regions
    if sheet_data['tables']:
        for table_idx, table in enumerate(sheet_data['tables']):
            table_text = _format_table_as_text(table)
            regions.append(Region(
                region_id=str(uuid.uuid4()),
                text=table_text,
                bbox=[0, table['start_row'] * 20, 100, table['end_row'] * 20],
                confidence=0.9,
                metadata={
                    "engine": "openpyxl",
                    "content_type": "table",
                    "sheet_name": sheet_data['name'],
                    "table_index": table_idx,
                    "table_stats": {
                        "rows": table['total_rows'],
                        "cols": table['total_cols'],
                        "range": f"Row {table['start_row']}-{table['end_row']}"
                    }
                }
            ))
    else:
        # Create a general data region for the entire sheet
        sheet_text = _format_sheet_as_text(sheet_data)
        regions.append(Region(
            region_id=str(uuid.uuid4()),
            text=sheet_text,
            bbox=[0, 0, 100, 100],
            confidence=0.8,
            metadata={
                "engine": "openpyxl",
                "content_type": "sheet_data",
                "sheet_name": sheet_data['name'],
                "sheet_stats": {
                    "cells": sheet_data['cell_count'],
                    "formulas": sheet_data['formula_count'],
                    "dimensions": sheet_data['dimensions']
                }
            }
        ))
    
    # Add formula region if there are formulas
    if sheet_data['has_formulas']:
        formula_text = _extract_formulas_as_text(sheet_data['rows'])
        if formula_text:
            regions.append(Region(
                region_id=str(uuid.uuid4()),
                text=formula_text,
                bbox=[0, 0, 50, 100],
                confidence=0.7,
                metadata={
                    "engine": "openpyxl",
                    "content_type": "formulas",
                    "sheet_name": sheet_data['name'],
                    "formula_count": sheet_data['formula_count']
                }
            ))
    
    return regions

def _format_table_as_text(table: Dict[str, Any]) -> str:
    """Convert table data to readable text format"""
    try:
        lines = []
        
        # Add table header
        lines.append(f"[TABLE: Rows {table['start_row']}-{table['end_row']}]")
        
        # Add headers
        headers = [h for h in table['headers'] if h.strip()]
        if headers:
            lines.append(' | '.join(headers))
            lines.append('-' * min(50, len(' | '.join(headers))))
        
        # Add data rows (limit to first 10 rows to avoid huge text blocks)
        for i, row in enumerate(table['data_rows'][:10]):
            row_text = ' | '.join(str(cell) for cell in row if str(cell).strip())
            if row_text.strip():
                lines.append(row_text)
        
        if len(table['data_rows']) > 10:
            lines.append(f"... ({len(table['data_rows']) - 10} more rows)")
        
        return '\n'.join(lines)
    except Exception:
        return f"[TABLE: {table.get('total_rows', 0)} rows, {table.get('total_cols', 0)} columns]"

def _format_sheet_as_text(sheet_data: Dict[str, Any]) -> str:
    """Convert sheet data to readable text format"""
    try:
        lines = []
        lines.append(f"[WORKSHEET: {sheet_data['name']}]")
        lines.append(f"Dimensions: {sheet_data['dimensions']}")
        lines.append(f"Cells: {sheet_data['cell_count']}, Formulas: {sheet_data['formula_count']}")
        lines.append("")
        
        # Add first few rows of data
        for row in sheet_data['rows'][:5]:
            row_cells = []
            for cell in row['cells']:
                if cell['value'] is not None:
                    row_cells.append(str(cell['display_value']))
            
            if row_cells:
                lines.append(f"Row {row['row_number']}: " + ' | '.join(row_cells))
        
        if len(sheet_data['rows']) > 5:
            lines.append(f"... ({len(sheet_data['rows']) - 5} more rows)")
        
        return '\n'.join(lines)
    except Exception:
        return f"[WORKSHEET: {sheet_data['name']} - {sheet_data['cell_count']} cells]"

def _extract_formulas_as_text(rows: List[Dict[str, Any]]) -> str:
    """Extract all formulas from the sheet as text"""
    try:
        formulas = []
        for row in rows:
            for cell in row['cells']:
                if cell['has_formula'] and cell['formula']:
                    formulas.append(f"{cell['coordinate']}: {cell['formula']}")
        
        if not formulas:
            return ""
        
        lines = ["[FORMULAS]"]
        lines.extend(formulas[:20])  # Limit to first 20 formulas
        
        if len(formulas) > 20:
            lines.append(f"... ({len(formulas) - 20} more formulas)")
        
        return '\n'.join(lines)
    except Exception:
        return "[FORMULAS: extraction failed]"